from __future__ import annotations

import re
from collections import defaultdict
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

# ------------------------------------------------------------
# Configuration
# ------------------------------------------------------------
# Script-relative root so the script works no matter where it is run from.
PROJECT_ROOT = Path(__file__).resolve().parent
OUTPUT_FILE = "quarto_reference_audit.xlsx"
OUTPUT_SHEET_LIMIT = 31

SKIP_DIR_NAMES = {
    ".git",
    ".quarto",
    "_site",
    "_book",
    "_freeze",
    "reference-audit",
    "__pycache__",
    ".ipynb_checkpoints",
    ".venv",
    "venv",
}

TEXT_FILE_SUFFIXES = {".qmd", ".md", ".rmd", ".Rmd"}
BIB_FILE_SUFFIXES = {".bib"}

# ------------------------------------------------------------
# Patterns
# ------------------------------------------------------------
# Quarto / Pandoc ids inside attribute blocks
ATTRIBUTE_ID_RE = re.compile(r"\{[^{}]*#([A-Za-z0-9][A-Za-z0-9:_-]*)[^{}]*\}")

# Executable cell labels, e.g. #| label: fig-my-plot
CELL_LABEL_RE = re.compile(r"^\s*#\|\s*label\s*:\s*([A-Za-z0-9][A-Za-z0-9:_-]*)\s*$", re.MULTILINE)

# Cross-refs or citations like @fig-foo, @smith2021
AT_REF_RE = re.compile(r"(?<![A-Za-z0-9_])@([A-Za-z0-9][A-Za-z0-9:_-]*)")

# Anchor references like [text](#sec-methods)
ANCHOR_REF_RE = re.compile(r"\]\(#([A-Za-z0-9][A-Za-z0-9:_-]*)\)")

# Standard markdown links: [text](target)
MARKDOWN_LINK_RE = re.compile(r"\[([^\]]*)\]\(([^)]+)\)")

# Autolinks: <https://example.com>
AUTOLINK_RE = re.compile(r"<(https?://[^>\s]+)>")

# YAML bibliography declarations
BIBLIOGRAPHY_LINE_RE = re.compile(r"^\s*bibliography\s*:\s*(.+?)\s*$", re.MULTILINE)

# BibTeX entry keys
BIB_ENTRY_RE = re.compile(r"^\s*@\s*([A-Za-z]+)\s*\{\s*([^,\s]+)\s*,", re.MULTILINE)

CROSSREF_PREFIXES = (
    "fig-",
    "tbl-",
    "eq-",
    "sec-",
    "lst-",
    "thm-",
    "lem-",
    "cor-",
    "prp-",
    "cnj-",
    "def-",
    "exm-",
    "exr-",
    "sol-",
    "rem-",
    "alg-",
    "apx-",
)

# ------------------------------------------------------------
# Helpers
# ------------------------------------------------------------
def safe_read_text(path: Path) -> str:
    for encoding in ("utf-8", "utf-8-sig", "latin-1"):
        try:
            return path.read_text(encoding=encoding)
        except UnicodeDecodeError:
            continue
    return path.read_text(encoding="utf-8", errors="replace")


def should_skip(path: Path) -> bool:
    return any(part in SKIP_DIR_NAMES for part in path.parts)


def iter_files(root: Path, suffixes: set[str]) -> list[Path]:
    return sorted(
        p for p in root.rglob("*") if p.is_file() and p.suffix in suffixes and not should_skip(p)
    )


def relative_str(path: Path, root: Path) -> str:
    return path.relative_to(root).as_posix()


def line_number(text: str, position: int) -> int:
    return text.count("\n", 0, position) + 1


def line_context(lines: list[str], idx0: int) -> str:
    if 0 <= idx0 < len(lines):
        if lines[idx0].strip():
            return lines[idx0].strip()
        for j in range(max(0, idx0 - 2), min(len(lines), idx0 + 3)):
            if lines[j].strip():
                return lines[j].strip()
    return ""


def infer_label_type(label: str) -> str:
    mapping = {
        "fig-": "Figure",
        "tbl-": "Table",
        "eq-": "Equation",
        "sec-": "Section",
        "lst-": "Listing",
        "thm-": "Theorem",
        "lem-": "Lemma",
        "cor-": "Corollary",
        "prp-": "Proposition",
        "cnj-": "Conjecture",
        "def-": "Definition",
        "exm-": "Example",
        "exr-": "Exercise",
        "sol-": "Solution",
        "rem-": "Remark",
        "alg-": "Algorithm",
        "apx-": "Appendix",
    }
    for prefix, kind in mapping.items():
        if label.startswith(prefix):
            return kind
    return "Other"


def is_crossref_key(key: str) -> bool:
    return key.startswith(CROSSREF_PREFIXES)


def clean_bibliography_value(raw: str) -> list[str]:
    raw = raw.strip()
    if not raw:
        return []
    if raw.endswith(".bib") and not raw.startswith("["):
        return [raw.strip().strip("\"'")]
    if raw.startswith("[") and raw.endswith("]"):
        inner = raw[1:-1].strip()
        if not inner:
            return []
        return [piece.strip().strip("\"'") for piece in inner.split(",") if piece.strip()]
    return []


def dedupe_rows(rows: Iterable[dict], key_fields: tuple[str, ...]) -> list[dict]:
    seen: set[tuple] = set()
    output: list[dict] = []
    for row in rows:
        key = tuple(row[field] for field in key_fields)
        if key not in seen:
            seen.add(key)
            output.append(row)
    return output


def classify_link_target(target: str) -> str:
    target = target.strip()
    lower = target.lower()
    if lower.startswith(("http://", "https://", "mailto:")):
        return "external"
    if target.startswith("#"):
        return "anchor"
    if lower.endswith((".qmd", ".md", ".html", ".htm", ".pdf", ".docx", ".xlsx", ".csv", ".png", ".jpg", ".jpeg", ".svg")):
        return "internal-file"
    return "other"


def normalize_link_target(target: str) -> str:
    return target.strip().split()[0]

# ------------------------------------------------------------
# Extractors
# ------------------------------------------------------------
def extract_labels(text: str, relpath: str) -> list[dict]:
    rows: list[dict] = []
    lines = text.splitlines()

    for match in ATTRIBUTE_ID_RE.finditer(text):
        label = match.group(1)
        lineno = line_number(text, match.start())
        rows.append(
            {
                "label": label,
                "type": infer_label_type(label),
                "file": relpath,
                "line": lineno,
                "source": "attribute-id",
                "context": line_context(lines, lineno - 1),
            }
        )

    for match in CELL_LABEL_RE.finditer(text):
        label = match.group(1)
        lineno = line_number(text, match.start())
        rows.append(
            {
                "label": label,
                "type": infer_label_type(label),
                "file": relpath,
                "line": lineno,
                "source": "cell-label",
                "context": line_context(lines, lineno - 1),
            }
        )

    return dedupe_rows(rows, ("label", "file", "line", "source"))


def extract_label_references(text: str, relpath: str) -> list[dict]:
    rows: list[dict] = []
    lines = text.splitlines()

    for match in AT_REF_RE.finditer(text):
        ref = match.group(1)
        if is_crossref_key(ref):
            lineno = line_number(text, match.start())
            rows.append(
                {
                    "reference": ref,
                    "type": infer_label_type(ref),
                    "file": relpath,
                    "line": lineno,
                    "syntax": "@id",
                    "context": line_context(lines, lineno - 1),
                }
            )

    for match in ANCHOR_REF_RE.finditer(text):
        ref = match.group(1)
        lineno = line_number(text, match.start())
        rows.append(
            {
                "reference": ref,
                "type": infer_label_type(ref),
                "file": relpath,
                "line": lineno,
                "syntax": "(#id)",
                "context": line_context(lines, lineno - 1),
            }
        )

    return dedupe_rows(rows, ("reference", "file", "line", "syntax"))


def extract_bib_citations(text: str, relpath: str) -> list[dict]:
    rows: list[dict] = []
    lines = text.splitlines()
    for match in AT_REF_RE.finditer(text):
        key = match.group(1)
        if not is_crossref_key(key):
            lineno = line_number(text, match.start())
            rows.append(
                {
                    "citation_key": key,
                    "file": relpath,
                    "line": lineno,
                    "context": line_context(lines, lineno - 1),
                }
            )
    return dedupe_rows(rows, ("citation_key", "file", "line"))


def extract_declared_bib_paths(text: str, qmd_path: Path) -> list[Path]:
    paths: list[Path] = []
    for match in BIBLIOGRAPHY_LINE_RE.finditer(text):
        raw = match.group(1)
        for item in clean_bibliography_value(raw):
            paths.append((qmd_path.parent / item).resolve())
    return paths


def extract_bib_entries(text: str, relpath: str) -> list[dict]:
    lines = text.splitlines()
    rows: list[dict] = []
    for match in BIB_ENTRY_RE.finditer(text):
        entry_type = match.group(1)
        bib_key = match.group(2)
        lineno = line_number(text, match.start())
        rows.append(
            {
                "bib_key": bib_key,
                "entry_type": entry_type,
                "file": relpath,
                "line": lineno,
                "context": line_context(lines, lineno - 1),
            }
        )
    return dedupe_rows(rows, ("bib_key", "file", "line"))


def extract_hyperlinks(text: str, relpath: str) -> list[dict]:
    rows: list[dict] = []
    lines = text.splitlines()

    for match in MARKDOWN_LINK_RE.finditer(text):
        link_text = match.group(1).strip()
        target = normalize_link_target(match.group(2))
        lineno = line_number(text, match.start())
        rows.append(
            {
                "link_text": link_text,
                "target": target,
                "link_type": classify_link_target(target),
                "file": relpath,
                "line": lineno,
                "context": line_context(lines, lineno - 1),
            }
        )

    for match in AUTOLINK_RE.finditer(text):
        target = normalize_link_target(match.group(1))
        lineno = line_number(text, match.start())
        rows.append(
            {
                "link_text": "",
                "target": target,
                "link_type": "external",
                "file": relpath,
                "line": lineno,
                "context": line_context(lines, lineno - 1),
            }
        )

    return dedupe_rows(rows, ("link_text", "target", "file", "line"))

# ------------------------------------------------------------
# Excel writer
# ------------------------------------------------------------
def autofit_worksheet(ws) -> None:
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for column_cells in ws.columns:
        letter = get_column_letter(column_cells[0].column)
        max_len = 0
        for cell in list(column_cells)[:200]:
            val = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(val))
        ws.column_dimensions[letter].width = min(max_len + 2, 80)
    for cell in ws[1]:
        cell.font = Font(bold=True)


def add_sheet(wb: Workbook, title: str, headers: list[str], rows: list[dict]) -> None:
    ws = wb.create_sheet(title=title[:OUTPUT_SHEET_LIMIT])
    ws.append(headers)
    for row in rows:
        ws.append([row.get(header, "") for header in headers])
    autofit_worksheet(ws)

# ------------------------------------------------------------
# Main
# ------------------------------------------------------------
def main() -> None:
    if not PROJECT_ROOT.exists():
        raise FileNotFoundError(f"Project root not found: {PROJECT_ROOT}")

    qmd_files = iter_files(PROJECT_ROOT, TEXT_FILE_SUFFIXES)
    bib_files = iter_files(PROJECT_ROOT, BIB_FILE_SUFFIXES)

    labels: list[dict] = []
    label_refs: list[dict] = []
    bib_citations: list[dict] = []
    bib_entries: list[dict] = []
    hyperlinks: list[dict] = []
    declared_bib_paths: set[Path] = set()

    for file_path in qmd_files:
        relpath = relative_str(file_path, PROJECT_ROOT)
        text = safe_read_text(file_path)
        labels.extend(extract_labels(text, relpath))
        label_refs.extend(extract_label_references(text, relpath))
        bib_citations.extend(extract_bib_citations(text, relpath))
        hyperlinks.extend(extract_hyperlinks(text, relpath))
        declared_bib_paths.update(extract_declared_bib_paths(text, file_path))

    for bib_file in bib_files:
        relpath = relative_str(bib_file, PROJECT_ROOT)
        text = safe_read_text(bib_file)
        bib_entries.extend(extract_bib_entries(text, relpath))

    labels = dedupe_rows(labels, ("label", "file", "line", "source"))
    label_refs = dedupe_rows(label_refs, ("reference", "file", "line", "syntax"))
    bib_citations = dedupe_rows(bib_citations, ("citation_key", "file", "line"))
    bib_entries = dedupe_rows(bib_entries, ("bib_key", "file", "line"))
    hyperlinks = dedupe_rows(hyperlinks, ("link_text", "target", "file", "line"))

    label_to_rows: dict[str, list[dict]] = defaultdict(list)
    for row in labels:
        label_to_rows[row["label"]].append(row)

    ref_to_rows: dict[str, list[dict]] = defaultdict(list)
    for row in label_refs:
        ref_to_rows[row["reference"]].append(row)

    bibkey_to_rows: dict[str, list[dict]] = defaultdict(list)
    for row in bib_entries:
        bibkey_to_rows[row["bib_key"]].append(row)

    citation_to_rows: dict[str, list[dict]] = defaultdict(list)
    for row in bib_citations:
        citation_to_rows[row["citation_key"]].append(row)

    all_label_names = set(label_to_rows)
    all_ref_names = set(ref_to_rows)
    all_bib_keys = set(bibkey_to_rows)

    duplicate_labels: list[dict] = []
    duplicate_label_summary: list[dict] = []
    for label, rows in sorted(label_to_rows.items()):
        if len(rows) > 1:
            duplicate_label_summary.append(
                {
                    "label": label,
                    "type": rows[0]["type"],
                    "count": len(rows),
                    "files": "; ".join(sorted({r['file'] for r in rows})),
                }
            )
            for row in rows:
                duplicate_labels.append(
                    {
                        "label": label,
                        "type": row["type"],
                        "file": row["file"],
                        "line": row["line"],
                        "source": row["source"],
                        "context": row["context"],
                        "count": len(rows),
                    }
                )

    missing_label_refs: list[dict] = []
    for reference, rows in sorted(ref_to_rows.items()):
        if reference not in all_label_names:
            for row in rows:
                missing_label_refs.append(
                    {
                        "reference": reference,
                        "type": row["type"],
                        "file": row["file"],
                        "line": row["line"],
                        "syntax": row["syntax"],
                        "context": row["context"],
                    }
                )

    unused_labels: list[dict] = []
    for label, rows in sorted(label_to_rows.items()):
        if label not in all_ref_names:
            for row in rows:
                unused_labels.append(
                    {
                        "label": label,
                        "type": row["type"],
                        "file": row["file"],
                        "line": row["line"],
                        "source": row["source"],
                        "context": row["context"],
                    }
                )

    duplicate_bib_keys: list[dict] = []
    duplicate_bib_summary: list[dict] = []
    for bib_key, rows in sorted(bibkey_to_rows.items()):
        if len(rows) > 1:
            duplicate_bib_summary.append(
                {
                    "bib_key": bib_key,
                    "count": len(rows),
                    "files": "; ".join(sorted({r['file'] for r in rows})),
                }
            )
            for row in rows:
                duplicate_bib_keys.append(
                    {
                        "bib_key": bib_key,
                        "entry_type": row["entry_type"],
                        "file": row["file"],
                        "line": row["line"],
                        "context": row["context"],
                        "count": len(rows),
                    }
                )

    missing_bib_citations: list[dict] = []
    for key, rows in sorted(citation_to_rows.items()):
        if key not in all_bib_keys:
            for row in rows:
                missing_bib_citations.append(
                    {
                        "citation_key": key,
                        "file": row["file"],
                        "line": row["line"],
                        "context": row["context"],
                    }
                )

    missing_bib_files: list[dict] = []
    for path in sorted(declared_bib_paths):
        if not path.exists():
            missing_bib_files.append({"declared_bib_file": str(path), "status": "missing on disk"})

    external_links = sorted((r for r in hyperlinks if r["link_type"] == "external"), key=lambda r: (r["target"], r["file"], r["line"]))
    internal_links = sorted((r for r in hyperlinks if r["link_type"] == "internal-file"), key=lambda r: (r["target"], r["file"], r["line"]))
    anchor_links = sorted((r for r in hyperlinks if r["link_type"] == "anchor"), key=lambda r: (r["target"], r["file"], r["line"]))
    other_links = sorted((r for r in hyperlinks if r["link_type"] == "other"), key=lambda r: (r["target"], r["file"], r["line"]))

    broken_internal_links: list[dict] = []
    for row in internal_links:
        source_file = PROJECT_ROOT / row["file"]
        target_path = normalize_link_target(row["target"]).split("#", 1)[0]
        resolved = (source_file.parent / target_path).resolve()
        if not resolved.exists():
            broken_internal_links.append(
                {
                    "link_text": row["link_text"],
                    "target": row["target"],
                    "resolved_path": str(resolved),
                    "file": row["file"],
                    "line": row["line"],
                    "context": row["context"],
                }
            )

    output_path = PROJECT_ROOT / OUTPUT_FILE
    wb = Workbook()
    default_sheet = wb.active
    wb.remove(default_sheet)

    summary_rows = [
        {"metric": "project_root", "value": str(PROJECT_ROOT)},
        {"metric": "qmd_files_scanned", "value": len(qmd_files)},
        {"metric": "bib_files_scanned", "value": len(bib_files)},
        {"metric": "labels_found", "value": len(labels)},
        {"metric": "label_references_found", "value": len(label_refs)},
        {"metric": "duplicate_label_rows", "value": len(duplicate_labels)},
        {"metric": "duplicate_label_groups", "value": len(duplicate_label_summary)},
        {"metric": "missing_label_reference_rows", "value": len(missing_label_refs)},
        {"metric": "unused_label_rows", "value": len(unused_labels)},
        {"metric": "bib_entries_found", "value": len(bib_entries)},
        {"metric": "bib_citations_found", "value": len(bib_citations)},
        {"metric": "duplicate_bib_key_rows", "value": len(duplicate_bib_keys)},
        {"metric": "duplicate_bib_key_groups", "value": len(duplicate_bib_summary)},
        {"metric": "missing_bib_citation_rows", "value": len(missing_bib_citations)},
        {"metric": "missing_declared_bib_files", "value": len(missing_bib_files)},
        {"metric": "hyperlinks_found", "value": len(hyperlinks)},
        {"metric": "external_links", "value": len(external_links)},
        {"metric": "internal_links", "value": len(internal_links)},
        {"metric": "anchor_links", "value": len(anchor_links)},
        {"metric": "other_links", "value": len(other_links)},
        {"metric": "broken_internal_links", "value": len(broken_internal_links)},
        {"metric": "output_file", "value": str(output_path)},
    ]

    add_sheet(wb, "summary", ["metric", "value"], summary_rows)
    add_sheet(wb, "all_labels", ["label", "type", "file", "line", "source", "context"], sorted(labels, key=lambda r: (r["type"], r["label"], r["file"], r["line"])))
    add_sheet(wb, "all_label_refs", ["reference", "type", "file", "line", "syntax", "context"], sorted(label_refs, key=lambda r: (r["type"], r["reference"], r["file"], r["line"])))
    add_sheet(wb, "duplicate_labels", ["label", "type", "file", "line", "source", "context", "count"], duplicate_labels)
    add_sheet(wb, "duplicate_label_groups", ["label", "type", "count", "files"], duplicate_label_summary)
    add_sheet(wb, "missing_label_refs", ["reference", "type", "file", "line", "syntax", "context"], missing_label_refs)
    add_sheet(wb, "unused_labels", ["label", "type", "file", "line", "source", "context"], unused_labels)
    add_sheet(wb, "bib_entries", ["bib_key", "entry_type", "file", "line", "context"], sorted(bib_entries, key=lambda r: (r["bib_key"], r["file"], r["line"])))
    add_sheet(wb, "bib_citations", ["citation_key", "file", "line", "context"], sorted(bib_citations, key=lambda r: (r["citation_key"], r["file"], r["line"])))
    add_sheet(wb, "duplicate_bib_keys", ["bib_key", "entry_type", "file", "line", "context", "count"], duplicate_bib_keys)
    add_sheet(wb, "duplicate_bib_groups", ["bib_key", "count", "files"], duplicate_bib_summary)
    add_sheet(wb, "missing_bib_cites", ["citation_key", "file", "line", "context"], missing_bib_citations)
    add_sheet(wb, "missing_bib_files", ["declared_bib_file", "status"], missing_bib_files)
    add_sheet(wb, "hyperlinks", ["link_text", "target", "link_type", "file", "line", "context"], sorted(hyperlinks, key=lambda r: (r["link_type"], r["target"], r["file"], r["line"])))
    add_sheet(wb, "external_links", ["link_text", "target", "link_type", "file", "line", "context"], external_links)
    add_sheet(wb, "internal_links", ["link_text", "target", "link_type", "file", "line", "context"], internal_links)
    add_sheet(wb, "anchor_links", ["link_text", "target", "link_type", "file", "line", "context"], anchor_links)
    add_sheet(wb, "other_links", ["link_text", "target", "link_type", "file", "line", "context"], other_links)
    add_sheet(wb, "broken_internal_links", ["link_text", "target", "resolved_path", "file", "line", "context"], broken_internal_links)

    wb.save(output_path)

    print("\nQuarto reference audit complete.")
    print(f"Project root: {PROJECT_ROOT}")
    print(f"QMD/Markdown files scanned: {len(qmd_files)}")
    print(f"Bib files scanned: {len(bib_files)}")
    print(f"Labels found: {len(labels)}")
    print(f"Label references found: {len(label_refs)}")
    print(f"Duplicate label groups: {len(duplicate_label_summary)}")
    print(f"Missing label references: {len(missing_label_refs)}")
    print(f"Unused labels: {len(unused_labels)}")
    print(f"Bibliography entries found: {len(bib_entries)}")
    print(f"Bibliography citations found: {len(bib_citations)}")
    print(f"Duplicate bibliography key groups: {len(duplicate_bib_summary)}")
    print(f"Missing bibliography citations: {len(missing_bib_citations)}")
    print(f"Missing declared bibliography files: {len(missing_bib_files)}")
    print(f"Hyperlinks found: {len(hyperlinks)}")
    print(f"Broken internal links: {len(broken_internal_links)}")
    print(f"Workbook written to: {output_path}\n")


if __name__ == "__main__":
    main()
