from __future__ import annotations

import re
from collections import defaultdict
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

# ------------------------------------------------------------
# Configuration
# ------------------------------------------------------------
# Script-relative root so the script works no matter where it is run from.
PROJECT_ROOT = Path(__file__).resolve().parent
OUTPUT_FILE = "quarto_reference_audit.xlsx"
OUTPUT_SHEET_LIMIT = 31

SKIP_DIR_NAMES = {
    ".git",
    ".quarto",
    "_site",
    "_book",
    "_freeze",
    "reference-audit",
    "__pycache__",
    ".ipynb_checkpoints",
    ".venv",
    "venv",
}

TEXT_FILE_SUFFIXES = {".qmd", ".md", ".rmd", ".Rmd"}
BIB_FILE_SUFFIXES = {".bib"}

# ------------------------------------------------------------
# Patterns
# ------------------------------------------------------------
ATTRIBUTE_ID_RE = re.compile(
    r"\{[^{}]*#([A-Za-z0-9][A-Za-z0-9:_-]*)[^{}]*\}"
)
CELL_LABEL_RE = re.compile(
    r"^\s*#\|\s*label\s*:\s*([A-Za-z0-9][A-Za-z0-9:_-]*)\s*$",
    re.MULTILINE,
)
AT_REF_RE = re.compile(r"(?<![A-Za-z0-9_])@([A-Za-z0-9][A-Za-z0-9:_-]*)")
ANCHOR_REF_RE = re.compile(r"\]\(#([A-Za-z0-9][A-Za-z0-9:_-]*)\)")
MARKDOWN_LINK_RE = re.compile(r"\[([^\]]*)\]\(([^)]+)\)")
AUTOLINK_RE = re.compile(r"<(https?://[^>\s]+)>")
BIBLIOGRAPHY_LINE_RE = re.compile(
    r"^\s*bibliography\s*:\s*(.+?)\s*$", re.MULTILINE
)
BIB_ENTRY_RE = re.compile(
    r"^\s*@\s*([A-Za-z]+)\s*\{\s*([^,\s]+)\s*,", re.MULTILINE
)
DISPLAY_EQUATION_RE = re.compile(
    r":::\s*\{#(?P<eq_id>eq-[A-Za-z0-9][A-Za-z0-9:_-]*)\}\s*"
    r"(?P<pre_body>.*?)"
    r"\$\$(?P<body>.*?)\$\$\s*"
    r":::",
    re.DOTALL,
)
EQ_CUSTOM_ASIDE_RE = re.compile(
    r"\[\[\s*\*\*(?P<custom_id>[A-Za-z0-9.:-]+)\*\*\s*\]\]\s*"
    r"\{[^}]*\.aside[^}]*\}"
)
EQ_CUSTOM_HTML_RE = re.compile(
    r"<p[^>]*>\s*\[?(?P<custom_id>[A-Za-z0-9.:-]+)\]?\s*</p>"
)
EQ_CUSTOM_TEXT_RE = re.compile(r"\\text\s*\{\s*\(([^(){}]+)\)\s*\}")
EQ_CUSTOM_TAG_RE = re.compile(r"\\tag\s*\{\s*([^{}]+?)\s*\}")

LABEL_TYPE_BY_PREFIX = {
    "fig-": "Figure",
    "tbl-": "Table",
    "eq-": "Equation",
    "sec-": "Section",
    "lst-": "Listing",
    "thm-": "Theorem",
    "lem-": "Lemma",
    "cor-": "Corollary",
    "prp-": "Proposition",
    "cnj-": "Conjecture",
    "def-": "Definition",
    "exm-": "Example",
    "exr-": "Exercise",
    "sol-": "Solution",
    "rem-": "Remark",
    "alg-": "Algorithm",
    "apx-": "Appendix",
}
CROSSREF_PREFIXES = tuple(LABEL_TYPE_BY_PREFIX)

DEDUPE_KEYS = {
    "labels": ("label", "file", "line", "source"),
    "label_refs": ("reference", "file", "line", "syntax"),
    "bib_citations": ("citation_key", "file", "line"),
    "bib_entries": ("bib_key", "file", "line"),
    "hyperlinks": ("link_text", "target", "file", "line"),
    "equations": ("eq_id", "file", "line"),
}

SHEET_SPECS = [
    (
        "all_labels",
        ["label", "type", "file", "line", "source", "context"],
        "labels",
        lambda r: (r["type"], r["label"], r["file"], r["line"]),
    ),
    (
        "all_label_refs",
        ["reference", "type", "file", "line", "syntax", "context"],
        "label_refs",
        lambda r: (r["type"], r["reference"], r["file"], r["line"]),
    ),
    (
        "duplicate_labels",
        ["label", "type", "file", "line", "source", "context", "count"],
        "duplicate_labels",
        None,
    ),
    (
        "duplicate_label_groups",
        ["label", "type", "count", "files"],
        "duplicate_label_summary",
        None,
    ),
    (
        "missing_label_refs",
        ["reference", "type", "file", "line", "syntax", "context"],
        "missing_label_refs",
        None,
    ),
    (
        "unused_labels",
        ["label", "type", "file", "line", "source", "context"],
        "unused_labels",
        None,
    ),
    (
        "bib_entries",
        ["bib_key", "entry_type", "file", "line", "context"],
        "bib_entries",
        lambda r: (r["bib_key"], r["file"], r["line"]),
    ),
    (
        "bib_citations",
        ["citation_key", "file", "line", "context"],
        "bib_citations",
        lambda r: (r["citation_key"], r["file"], r["line"]),
    ),
    (
        "duplicate_bib_keys",
        ["bib_key", "entry_type", "file", "line", "context", "count"],
        "duplicate_bib_keys",
        None,
    ),
    (
        "duplicate_bib_groups",
        ["bib_key", "count", "files"],
        "duplicate_bib_summary",
        None,
    ),
    (
        "missing_bib_cites",
        ["citation_key", "file", "line", "context"],
        "missing_bib_citations",
        None,
    ),
    (
        "missing_bib_files",
        ["declared_bib_file", "status"],
        "missing_bib_files",
        None,
    ),
    (
        "hyperlinks",
        ["link_text", "target", "link_type", "file", "line", "context"],
        "hyperlinks",
        lambda r: (r["link_type"], r["target"], r["file"], r["line"]),
    ),
    (
        "external_links",
        ["link_text", "target", "link_type", "file", "line", "context"],
        "external_links",
        None,
    ),
    (
        "internal_links",
        ["link_text", "target", "link_type", "file", "line", "context"],
        "internal_links",
        None,
    ),
    (
        "anchor_links",
        ["link_text", "target", "link_type", "file", "line", "context"],
        "anchor_links",
        None,
    ),
    (
        "other_links",
        ["link_text", "target", "link_type", "file", "line", "context"],
        "other_links",
        None,
    ),
    (
        "broken_internal_links",
        ["link_text", "target", "resolved_path", "file", "line", "context"],
        "broken_internal_links",
        None,
    ),
    (
        "equations",
        [
            "eq_id",
            "custom_id",
            "expected_custom_id",
            "matches_expected",
            "has_custom_id",
            "file",
            "line",
            "equation_body",
            "context",
        ],
        "equations",
        lambda r: (r["file"], r["line"], r["eq_id"]),
    ),
    (
        "missing_eq_custom_ids",
        ["eq_id", "expected_custom_id", "file", "line", "equation_body", "context"],
        "missing_equation_custom_ids",
        None,
    ),
    (
        "equation_id_mismatches",
        [
            "eq_id",
            "custom_id",
            "expected_custom_id",
            "file",
            "line",
            "equation_body",
            "context",
        ],
        "equation_id_mismatches",
        None,
    ),
    (
        "dup_eq_custom_ids",
        ["custom_id", "eq_id", "file", "line", "equation_body", "context", "count"],
        "duplicate_equation_custom_ids",
        None,
    ),
    (
        "dup_eq_custom_groups",
        ["custom_id", "count", "files"],
        "duplicate_equation_custom_groups",
        None,
    ),
]

PRINT_METRICS = [
    ("Project root", "project_root"),
    ("QMD/Markdown files scanned", "qmd_files_scanned"),
    ("Bib files scanned", "bib_files_scanned"),
    ("Labels found", "labels_found"),
    ("Label references found", "label_references_found"),
    ("Duplicate label groups", "duplicate_label_groups"),
    ("Missing label references", "missing_label_reference_rows"),
    ("Unused labels", "unused_label_rows"),
    ("Bibliography entries found", "bib_entries_found"),
    ("Bibliography citations found", "bib_citations_found"),
    ("Duplicate bibliography key groups", "duplicate_bib_key_groups"),
    ("Missing bibliography citations", "missing_bib_citation_rows"),
    ("Missing declared bibliography files", "missing_declared_bib_files"),
    ("Hyperlinks found", "hyperlinks_found"),
    ("Broken internal links", "broken_internal_links"),
    ("Workbook written to", "output_file"),
]

# ------------------------------------------------------------
# General helpers
# ------------------------------------------------------------
def safe_read_text(path: Path) -> str:
    for encoding in ("utf-8", "utf-8-sig", "latin-1"):
        try:
            return path.read_text(encoding=encoding)
        except UnicodeDecodeError:
            continue
    return path.read_text(encoding="utf-8", errors="replace")


def should_skip(path: Path) -> bool:
    return any(part in SKIP_DIR_NAMES for part in path.parts)


def iter_files(root: Path, suffixes: set[str]) -> list[Path]:
    return sorted(
        p
        for p in root.rglob("*")
        if p.is_file() and p.suffix in suffixes and not should_skip(p)
    )


def relative_str(path: Path, root: Path = PROJECT_ROOT) -> str:
    return path.relative_to(root).as_posix()


def line_number(text: str, position: int) -> int:
    return text.count("\n", 0, position) + 1


def line_context(lines: list[str], idx0: int) -> str:
    if 0 <= idx0 < len(lines) and lines[idx0].strip():
        return lines[idx0].strip()

    start = max(0, idx0 - 2)
    stop = min(len(lines), idx0 + 3)
    for idx in range(start, stop):
        if lines[idx].strip():
            return lines[idx].strip()
    return ""


def dedupe_rows(rows: Iterable[dict], key_fields: tuple[str, ...]) -> list[dict]:
    seen: set[tuple] = set()
    output: list[dict] = []
    for row in rows:
        key = tuple(row[field] for field in key_fields)
        if key in seen:
            continue
        seen.add(key)
        output.append(row)
    return output


def dedupe_audit(audit: dict) -> dict:
    for section, key_fields in DEDUPE_KEYS.items():
        audit[section] = dedupe_rows(audit[section], key_fields)
    return audit


def group_by(rows: Iterable[dict], key: str) -> dict[str, list[dict]]:
    grouped: dict[str, list[dict]] = defaultdict(list)
    for row in rows:
        grouped[row[key]].append(row)
    return grouped


def infer_label_type(label: str) -> str:
    for prefix, kind in LABEL_TYPE_BY_PREFIX.items():
        if label.startswith(prefix):
            return kind
    return "Other"


def is_crossref_key(key: str) -> bool:
    return key.startswith(CROSSREF_PREFIXES)


def clean_bibliography_value(raw: str) -> list[str]:
    raw = raw.strip()
    if not raw:
        return []
    if raw.endswith(".bib") and not raw.startswith("["):
        return [raw.strip().strip("\"'")]
    if raw.startswith("[") and raw.endswith("]"):
        items = raw[1:-1].strip().split(",")
        return [item.strip().strip("\"'") for item in items if item.strip()]
    return []


def classify_link_target(target: str) -> str:
    target = target.strip()
    lower = target.lower()
    if lower.startswith(("http://", "https://", "mailto:")):
        return "external"
    if target.startswith("#"):
        return "anchor"
    file_suffixes = (
        ".qmd",
        ".md",
        ".html",
        ".htm",
        ".pdf",
        ".docx",
        ".xlsx",
        ".csv",
        ".png",
        ".jpg",
        ".jpeg",
        ".svg",
    )
    if lower.endswith(file_suffixes):
        return "internal-file"
    return "other"


def normalize_link_target(target: str) -> str:
    return target.strip().split()[0]


def normalize_custom_equation_id(raw: str) -> str:
    raw = raw.strip()
    wrapper_re = re.compile(r"\\(?:textbf|mathrm|mathbf|textrm)\s*\{([^{}]+)\}")
    while True:
        match = wrapper_re.fullmatch(raw)
        if not match:
            break
        raw = match.group(1).strip()
    return raw.strip("()[]{} ")


def expected_custom_equation_id(eq_id: str) -> str:
    core = eq_id[3:] if eq_id.startswith("eq-") else eq_id
    return core.replace("-", ".").upper()

# ------------------------------------------------------------
# Extractors
# ------------------------------------------------------------
def make_label_row(label: str, relpath: str, lineno: int, source: str, lines: list[str]) -> dict:
    return {
        "label": label,
        "type": infer_label_type(label),
        "file": relpath,
        "line": lineno,
        "source": source,
        "context": line_context(lines, lineno - 1),
    }


def extract_labels(text: str, relpath: str) -> list[dict]:
    rows: list[dict] = []
    lines = text.splitlines()

    for match in ATTRIBUTE_ID_RE.finditer(text):
        lineno = line_number(text, match.start())
        rows.append(make_label_row(match.group(1), relpath, lineno, "attribute-id", lines))

    for match in CELL_LABEL_RE.finditer(text):
        lineno = line_number(text, match.start())
        rows.append(make_label_row(match.group(1), relpath, lineno, "cell-label", lines))

    return rows


def extract_label_references(text: str, relpath: str) -> list[dict]:
    rows: list[dict] = []
    lines = text.splitlines()

    for match in AT_REF_RE.finditer(text):
        ref = match.group(1)
        if is_crossref_key(ref):
            lineno = line_number(text, match.start())
            rows.append(
                {
                    "reference": ref,
                    "type": infer_label_type(ref),
                    "file": relpath,
                    "line": lineno,
                    "syntax": "@id",
                    "context": line_context(lines, lineno - 1),
                }
            )

    for match in ANCHOR_REF_RE.finditer(text):
        ref = match.group(1)
        lineno = line_number(text, match.start())
        rows.append(
            {
                "reference": ref,
                "type": infer_label_type(ref),
                "file": relpath,
                "line": lineno,
                "syntax": "(#id)",
                "context": line_context(lines, lineno - 1),
            }
        )

    return rows


def extract_bib_citations(text: str, relpath: str) -> list[dict]:
    rows: list[dict] = []
    lines = text.splitlines()
    for match in AT_REF_RE.finditer(text):
        key = match.group(1)
        if is_crossref_key(key):
            continue
        lineno = line_number(text, match.start())
        rows.append(
            {
                "citation_key": key,
                "file": relpath,
                "line": lineno,
                "context": line_context(lines, lineno - 1),
            }
        )
    return rows


def extract_declared_bib_paths(text: str, qmd_path: Path) -> set[Path]:
    paths: set[Path] = set()
    for match in BIBLIOGRAPHY_LINE_RE.finditer(text):
        for item in clean_bibliography_value(match.group(1)):
            paths.add((qmd_path.parent / item).resolve())
    return paths


def extract_bib_entries(text: str, relpath: str) -> list[dict]:
    rows: list[dict] = []
    lines = text.splitlines()
    for match in BIB_ENTRY_RE.finditer(text):
        lineno = line_number(text, match.start())
        rows.append(
            {
                "bib_key": match.group(2),
                "entry_type": match.group(1),
                "file": relpath,
                "line": lineno,
                "context": line_context(lines, lineno - 1),
            }
        )
    return rows


def extract_hyperlinks(text: str, relpath: str) -> list[dict]:
    rows: list[dict] = []
    lines = text.splitlines()

    for match in MARKDOWN_LINK_RE.finditer(text):
        target = normalize_link_target(match.group(2))
        lineno = line_number(text, match.start())
        rows.append(
            {
                "link_text": match.group(1).strip(),
                "target": target,
                "link_type": classify_link_target(target),
                "file": relpath,
                "line": lineno,
                "context": line_context(lines, lineno - 1),
            }
        )

    for match in AUTOLINK_RE.finditer(text):
        target = normalize_link_target(match.group(1))
        lineno = line_number(text, match.start())
        rows.append(
            {
                "link_text": "",
                "target": target,
                "link_type": "external",
                "file": relpath,
                "line": lineno,
                "context": line_context(lines, lineno - 1),
            }
        )

    return rows


def find_custom_equation_id(pre_body: str, eq_body: str) -> str:
    search_order = [
        (EQ_CUSTOM_ASIDE_RE, pre_body, "custom_id"),
        (EQ_CUSTOM_HTML_RE, pre_body, "custom_id"),
        (EQ_CUSTOM_TEXT_RE, eq_body, 1),
        (EQ_CUSTOM_TAG_RE, eq_body, 1),
    ]
    for pattern, source, group in search_order:
        match = pattern.search(source)
        if match:
            return normalize_custom_equation_id(match.group(group))
    return ""


def extract_equation_registry(text: str, relpath: str) -> list[dict]:
    rows: list[dict] = []
    lines = text.splitlines()

    for match in DISPLAY_EQUATION_RE.finditer(text):
        eq_body = match.group("body").strip()
        eq_id = match.group("eq_id")
        lineno = line_number(text, match.start())
        custom_id = find_custom_equation_id(match.group("pre_body"), eq_body)
        expected_id = expected_custom_equation_id(eq_id)

        rows.append(
            {
                "eq_id": eq_id,
                "custom_id": custom_id,
                "expected_custom_id": expected_id,
                "matches_expected": "YES" if custom_id == expected_id else "NO",
                "has_custom_id": "YES" if custom_id else "NO",
                "file": relpath,
                "line": lineno,
                "equation_body": " ".join(eq_body.split()),
                "context": line_context(lines, lineno - 1),
            }
        )

    return rows

# ------------------------------------------------------------
# Collection and analysis
# ------------------------------------------------------------
def empty_audit() -> dict:
    return {
        "labels": [],
        "label_refs": [],
        "bib_citations": [],
        "bib_entries": [],
        "hyperlinks": [],
        "equations": [],
        "declared_bib_paths": set(),
    }


def collect_qmd_data(audit: dict, file_path: Path) -> None:
    relpath = relative_str(file_path)
    text = safe_read_text(file_path)

    audit["labels"].extend(extract_labels(text, relpath))
    audit["label_refs"].extend(extract_label_references(text, relpath))
    audit["bib_citations"].extend(extract_bib_citations(text, relpath))
    audit["hyperlinks"].extend(extract_hyperlinks(text, relpath))
    audit["equations"].extend(extract_equation_registry(text, relpath))
    audit["declared_bib_paths"].update(extract_declared_bib_paths(text, file_path))


def collect_bib_data(audit: dict, bib_file: Path) -> None:
    relpath = relative_str(bib_file)
    text = safe_read_text(bib_file)
    audit["bib_entries"].extend(extract_bib_entries(text, relpath))


def collect_project_audit(qmd_files: list[Path], bib_files: list[Path]) -> dict:
    audit = empty_audit()
    for file_path in qmd_files:
        collect_qmd_data(audit, file_path)
    for bib_file in bib_files:
        collect_bib_data(audit, bib_file)
    return dedupe_audit(audit)


def duplicate_groups(
    grouped_rows: dict[str, list[dict]],
    row_builder,
    summary_builder,
) -> tuple[list[dict], list[dict]]:
    duplicate_rows: list[dict] = []
    summary_rows: list[dict] = []

    for key, rows in sorted(grouped_rows.items()):
        if len(rows) <= 1:
            continue
        summary_rows.append(summary_builder(key, rows))
        duplicate_rows.extend(row_builder(key, row, len(rows)) for row in rows)

    return duplicate_rows, summary_rows


def find_missing_references(
    grouped_refs: dict[str, list[dict]],
    existing_keys: set[str],
    key_name: str,
) -> list[dict]:
    missing: list[dict] = []
    for key, rows in sorted(grouped_refs.items()):
        if key in existing_keys:
            continue
        for row in rows:
            output = {key_name: key}
            output.update(row)
            missing.append(output)
    return missing


def find_unused_labels(label_to_rows: dict[str, list[dict]], used_refs: set[str]) -> list[dict]:
    unused: list[dict] = []
    fields = ["label", "type", "file", "line", "source", "context"]
    for label, rows in sorted(label_to_rows.items()):
        if label in used_refs:
            continue
        unused.extend({field: row[field] for field in fields} for row in rows)
    return unused


def split_links_by_type(hyperlinks: list[dict]) -> dict[str, list[dict]]:
    links: dict[str, list[dict]] = {
        "external_links": [],
        "internal_links": [],
        "anchor_links": [],
        "other_links": [],
    }
    key_by_type = {
        "external": "external_links",
        "internal-file": "internal_links",
        "anchor": "anchor_links",
        "other": "other_links",
    }

    for row in hyperlinks:
        links[key_by_type[row["link_type"]]].append(row)

    for key, rows in links.items():
        links[key] = sorted(rows, key=lambda r: (r["target"], r["file"], r["line"]))

    return links


def find_broken_internal_links(internal_links: list[dict]) -> list[dict]:
    broken: list[dict] = []
    for row in internal_links:
        source_file = PROJECT_ROOT / row["file"]
        target_path = normalize_link_target(row["target"]).split("#", 1)[0]
        resolved = (source_file.parent / target_path).resolve()
        if resolved.exists():
            continue
        broken.append(
            {
                "link_text": row["link_text"],
                "target": row["target"],
                "resolved_path": str(resolved),
                "file": row["file"],
                "line": row["line"],
                "context": row["context"],
            }
        )
    return broken


def analyze_equations(equations: list[dict]) -> dict:
    custom_id_to_rows = group_by(
        (row for row in equations if row["custom_id"]), "custom_id"
    )

    missing_custom_ids = []
    mismatches = []
    for row in equations:
        if not row["custom_id"]:
            missing_custom_ids.append(
                pick_fields(
                    row,
                    [
                        "eq_id",
                        "expected_custom_id",
                        "file",
                        "line",
                        "equation_body",
                        "context",
                    ],
                )
            )
        elif row["custom_id"] != row["expected_custom_id"]:
            mismatches.append(
                pick_fields(
                    row,
                    [
                        "eq_id",
                        "custom_id",
                        "expected_custom_id",
                        "file",
                        "line",
                        "equation_body",
                        "context",
                    ],
                )
            )

    duplicate_rows, duplicate_groups_rows = duplicate_groups(
        custom_id_to_rows,
        build_duplicate_equation_row,
        build_duplicate_equation_summary,
    )

    return {
        "missing_equation_custom_ids": missing_custom_ids,
        "equation_id_mismatches": mismatches,
        "duplicate_equation_custom_ids": duplicate_rows,
        "duplicate_equation_custom_groups": duplicate_groups_rows,
    }


def pick_fields(row: dict, fields: list[str]) -> dict:
    return {field: row[field] for field in fields}


def build_duplicate_label_row(label: str, row: dict, count: int) -> dict:
    output = pick_fields(row, ["type", "file", "line", "source", "context"])
    output.update({"label": label, "count": count})
    return output


def build_duplicate_label_summary(label: str, rows: list[dict]) -> dict:
    return {
        "label": label,
        "type": rows[0]["type"],
        "count": len(rows),
        "files": "; ".join(sorted({row["file"] for row in rows})),
    }


def build_duplicate_bib_row(bib_key: str, row: dict, count: int) -> dict:
    output = pick_fields(row, ["entry_type", "file", "line", "context"])
    output.update({"bib_key": bib_key, "count": count})
    return output


def build_duplicate_bib_summary(bib_key: str, rows: list[dict]) -> dict:
    return {
        "bib_key": bib_key,
        "count": len(rows),
        "files": "; ".join(sorted({row["file"] for row in rows})),
    }


def build_duplicate_equation_row(custom_id: str, row: dict, count: int) -> dict:
    output = pick_fields(row, ["eq_id", "file", "line", "equation_body", "context"])
    output.update({"custom_id": custom_id, "count": count})
    return output


def build_duplicate_equation_summary(custom_id: str, rows: list[dict]) -> dict:
    return {
        "custom_id": custom_id,
        "count": len(rows),
        "files": "; ".join(sorted({row["file"] for row in rows})),
    }


def analyze_audit(audit: dict) -> dict:
    label_to_rows = group_by(audit["labels"], "label")
    ref_to_rows = group_by(audit["label_refs"], "reference")
    bibkey_to_rows = group_by(audit["bib_entries"], "bib_key")
    citation_to_rows = group_by(audit["bib_citations"], "citation_key")

    duplicate_labels, duplicate_label_summary = duplicate_groups(
        label_to_rows,
        build_duplicate_label_row,
        build_duplicate_label_summary,
    )
    duplicate_bib_keys, duplicate_bib_summary = duplicate_groups(
        bibkey_to_rows,
        build_duplicate_bib_row,
        build_duplicate_bib_summary,
    )

    links = split_links_by_type(audit["hyperlinks"])
    analysis = {
        "duplicate_labels": duplicate_labels,
        "duplicate_label_summary": duplicate_label_summary,
        "missing_label_refs": find_missing_references(
            ref_to_rows, set(label_to_rows), "reference"
        ),
        "unused_labels": find_unused_labels(label_to_rows, set(ref_to_rows)),
        "duplicate_bib_keys": duplicate_bib_keys,
        "duplicate_bib_summary": duplicate_bib_summary,
        "missing_bib_citations": find_missing_references(
            citation_to_rows, set(bibkey_to_rows), "citation_key"
        ),
        "missing_bib_files": [
            {"declared_bib_file": str(path), "status": "missing on disk"}
            for path in sorted(audit["declared_bib_paths"])
            if not path.exists()
        ],
        **links,
    }
    analysis["broken_internal_links"] = find_broken_internal_links(
        analysis["internal_links"]
    )
    analysis.update(analyze_equations(audit["equations"]))
    return analysis

# ------------------------------------------------------------
# Excel writer and reporting
# ------------------------------------------------------------
def autofit_worksheet(ws) -> None:
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for column_cells in ws.columns:
        letter = get_column_letter(column_cells[0].column)
        max_len = max(
            len("" if cell.value is None else str(cell.value))
            for cell in list(column_cells)[:200]
        )
        ws.column_dimensions[letter].width = min(max_len + 2, 80)
    for cell in ws[1]:
        cell.font = Font(bold=True)


def add_sheet(wb: Workbook, title: str, headers: list[str], rows: list[dict]) -> None:
    ws = wb.create_sheet(title=title[:OUTPUT_SHEET_LIMIT])
    ws.append(headers)
    for row in rows:
        ws.append([row.get(header, "") for header in headers])
    autofit_worksheet(ws)


def sort_rows(rows: list[dict], sort_key) -> list[dict]:
    return sorted(rows, key=sort_key) if sort_key else rows


def build_summary_rows(
    qmd_files: list[Path],
    bib_files: list[Path],
    audit: dict,
    analysis: dict,
    output_path: Path,
) -> list[dict]:
    metrics = {
        "project_root": str(PROJECT_ROOT),
        "qmd_files_scanned": len(qmd_files),
        "bib_files_scanned": len(bib_files),
        "labels_found": len(audit["labels"]),
        "label_references_found": len(audit["label_refs"]),
        "duplicate_label_rows": len(analysis["duplicate_labels"]),
        "duplicate_label_groups": len(analysis["duplicate_label_summary"]),
        "missing_label_reference_rows": len(analysis["missing_label_refs"]),
        "unused_label_rows": len(analysis["unused_labels"]),
        "bib_entries_found": len(audit["bib_entries"]),
        "bib_citations_found": len(audit["bib_citations"]),
        "duplicate_bib_key_rows": len(analysis["duplicate_bib_keys"]),
        "duplicate_bib_key_groups": len(analysis["duplicate_bib_summary"]),
        "missing_bib_citation_rows": len(analysis["missing_bib_citations"]),
        "missing_declared_bib_files": len(analysis["missing_bib_files"]),
        "hyperlinks_found": len(audit["hyperlinks"]),
        "external_links": len(analysis["external_links"]),
        "internal_links": len(analysis["internal_links"]),
        "anchor_links": len(analysis["anchor_links"]),
        "other_links": len(analysis["other_links"]),
        "broken_internal_links": len(analysis["broken_internal_links"]),
        "output_file": str(output_path),
        "equations_found": len(audit["equations"]),
        "missing_equation_custom_ids": len(analysis["missing_equation_custom_ids"]),
        "equation_id_mismatches": len(analysis["equation_id_mismatches"]),
        "duplicate_equation_custom_rows": len(
            analysis["duplicate_equation_custom_ids"]
        ),
        "duplicate_equation_custom_groups": len(
            analysis["duplicate_equation_custom_groups"]
        ),
    }
    return [{"metric": key, "value": value} for key, value in metrics.items()]


def write_workbook(
    qmd_files: list[Path],
    bib_files: list[Path],
    audit: dict,
    analysis: dict,
) -> tuple[Path, list[dict]]:
    output_path = PROJECT_ROOT / OUTPUT_FILE
    workbook_data = {**audit, **analysis}
    summary_rows = build_summary_rows(
        qmd_files, bib_files, audit, analysis, output_path
    )

    wb = Workbook()
    wb.remove(wb.active)
    add_sheet(wb, "summary", ["metric", "value"], summary_rows)

    for title, headers, data_key, sort_key in SHEET_SPECS:
        add_sheet(wb, title, headers, sort_rows(workbook_data[data_key], sort_key))

    wb.save(output_path)
    return output_path, summary_rows


def print_summary(summary_rows: list[dict]) -> None:
    summary = {row["metric"]: row["value"] for row in summary_rows}
    print("\nQuarto reference audit complete.")
    for label, metric in PRINT_METRICS:
        print(f"{label}: {summary[metric]}")
    print()

# ------------------------------------------------------------
# Main
# ------------------------------------------------------------
def validate_project_root() -> None:
    if not PROJECT_ROOT.exists():
        raise FileNotFoundError(f"Project root not found: {PROJECT_ROOT}")


def main() -> None:
    validate_project_root()
    qmd_files = iter_files(PROJECT_ROOT, TEXT_FILE_SUFFIXES)
    bib_files = iter_files(PROJECT_ROOT, BIB_FILE_SUFFIXES)
    audit = collect_project_audit(qmd_files, bib_files)
    analysis = analyze_audit(audit)
    _, summary_rows = write_workbook(qmd_files, bib_files, audit, analysis)
    print_summary(summary_rows)


if __name__ == "__main__":
    main()
